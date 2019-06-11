import openpyxl
import re
import os
import progressbar

print("开始处理文件夹中所有 .xlsx 文件……\r\n")
for excelFile in os.listdir('.'):
    if os.path.splitext(excelFile)[-1] == '.xlsx':
        print('读取 {} 文件……'.format(excelFile))
        wb = openpyxl.load_workbook(excelFile)
        #sheetnames = wb.get_sheet_names()
        sheet = wb[wb.sheetnames[0]]
        print('\r\n提取 {} 中的ip：'.format(excelFile))

        # 加上进度条显示
        for rowNum in progressbar.progressbar(range(2, sheet.max_row)):
            # print(sheet['F' + str(rowNum)].value)
            ipRex = re.compile(r'(\d{1,3}\.){1,3}\d{1,3}')
            ip = ipRex.search(sheet['F'+str(rowNum)].value)
            # print(ip.group())
            sheet.cell(row=rowNum, column=7).value = ip.group()
        print('保存 {} 文件……'.format(excelFile))
        wb.save(excelFile)

print("\r\n文件夹中所有 .xlsx 文件处理完成。")